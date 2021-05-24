# !/usr/python
# -*- coding:utf-8 -*-
'''
* rect
* |
* |___ sheet
*     |
*     |___table1
*     |
*     |___table2
*     |
*     ...
*
* table (bRow, bCol, eRow, eCol)
'''

import openpyxl
from pathlib import Path
from xlc import xmlExcel

class excel:
    def __init__(self, excelfile, formatfile):
        try:
            self.is_exist = Path(excelfile).exists()
            self.excel = None
            self.excelfile = excelfile
            self.format = xmlExcel.xmlExcel.parseXML(formatfile)
            if self.format != None:
                self.excel = self.format.createExcelTp()
                self.initExcel()
        except  Exception as e:
            print(e)
            self.format = None

    def initExcel(self):
        self.rect = {}
        if not self.is_exist:
            wb = openpyxl.Workbook()
            names = wb.sheetnames.copy()
            for sheet in self.excel:
                wb.create_sheet(sheet[0])
                self.drawHeader(sheet,wb[sheet[0]])
            for n in names:
                ws = wb[n]
                wb.remove(ws)
            wb.save(self.excelfile)
        else:
            wb = openpyxl.load_workbook(self.excelfile)
        wb.close()

    def drawHeader(self, sheet, ws):
        rect = {}
        row = 1
        col = 1
        titleRow = 1
        for table in sheet[2]:
            bRow = row
            bCol = col
            titleRow = row
            cell = ws.cell(titleRow, col)
            cell.value = table[0]
            row = row + 1
            for header in table[2]:
                cell = ws.cell(row, col)
                col = col + 1
                cell.value = header[0]
            ws.merge_cells(start_row=titleRow, start_column=bCol,
                            end_row=titleRow, end_column=col-1)
            rect[table[0]] = (bRow, bCol, row, col)
            row = row + 1
        self.rect[sheet[0]] = rect

    def addRow(self, sheetname, tablename, datas):
        wb = openpyxl.load_workbook(self.excelfile)
        if not wb: self.initExcel()
        ws = wb[sheetname]
        if not ws: return
        if tablename not in self.rect[sheetname].keys(): return
        row = self.rect[sheetname][tablename][2] + 1
        col = self.rect[sheetname][tablename][1]
        for data in datas:
            cell = ws.cell(row, col)
            cell.value = data
            col = col + 1
        self.rect[sheetname][tablename] = (self.rect[sheetname][tablename][0],
                                        self.rect[sheetname][tablename][1], row,
                                        self.rect[sheetname][tablename][3])
        wb.save(self.excelfile)
        wb.close()

    def findRow(self, sheetname, tablename, header, value):
        wb = openpyxl.load_workbook(self.excelfile)
        if not wb: return -1
        ws = wb[sheetname]
        if not ws: return -1
        if tablename not in self.rect[sheetname].keys(): return -1
        headerRow = self.rect[sheetname][tablename][0] + 1
        targetCol = self.rect[sheetname][tablename][1]
        for i in range(1,100):
            cell = ws.cell(headerRow, i)
            if header == cell.value:
                targetCol = i
                break
        beginRow = headerRow + 1
        endRow = self.rect[sheetname][tablename][2]
        for i in range(beginRow, endRow):
            cell = ws.cell(i, targetCol)
            if cell.value == value:
                wb.close()
                return i
        wb.close()
        return -1
    
    def findCol(self, sheetname, tablename, header):
        wb = openpyxl.load_workbook(self.excelfile)
        if not wb: return -1
        ws = wb[sheetname]
        if not ws: return -1
        if tablename not in self.rect[sheetname].keys(): return -1
        headerRow = self.rect[sheetname][tablename][0] + 1
        targetCol = self.rect[sheetname][tablename][1]
        for i in range(1,100):
            cell = ws.cell(headerRow, i)
            if header == cell.value:
                targetCol = i
                break
        wb.close()
        return targetCol

    def __findCell(self, sheetname, tablename, header, row):
        wb = openpyxl.load_workbook(self.excelfile)
        targetCol = self.findCol(sheetname, tablename, header)
        if targetCol == -1: return None
        return [wb, wb.cell(row, targetCol)]

    def setCell(self, sheetname, tablename, header, row, value):
        wb = openpyxl.load_workbook(self.excelfile)
        if not wb: return False
        ws = wb[sheetname]
        if not ws: return False
        if tablename not in self.rect[sheetname].keys(): return False
        headerRow = self.rect[sheetname][tablename][0] + 1
        targetCol = self.rect[sheetname][tablename][1]
        for i in range(1,100):
            cell = ws.cell(headerRow, i)
            if header == cell.value:
                targetCol = i
                break
        ws.cell(row, targetCol).value = value
        wb.save(self.excelfile)
        wb.close()
        return True

    def setDatas(self, sheetname, tablename, row, col, datas):
        wb = openpyxl.load_workbook(self.excelfile)
        if not wb: return False
        ws = wb[sheetname]
        if not ws: return False
        if tablename not in self.rect[sheetname].keys(): return False
        for c in range(len(datas)):
            cell = ws.cell(row, c + col)
            cell.value = datas[c]
        wb.save(self.excelfile)
        wb.close()
        return True